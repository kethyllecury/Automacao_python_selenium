import os
import time
import pandas as pd
from dotenv import load_dotenv
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from datetime import datetime, timedelta
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.edge.service import Service
from webdriver_manager.microsoft import EdgeChromiumDriverManager

load_dotenv("cripto.env")

dia_anterior = datetime.now() - timedelta(days=1)
data_anterior = dia_anterior.strftime("%d-%m-%Y")

hoje = datetime.today()

mes_atual = datetime.today().strftime("%m-%Y")
mes_anterior = hoje.replace(day=1) - timedelta(days=1)  # Vai para o último dia do mês anterior
mes_anterior = mes_anterior.strftime("%m-%Y")

data_atual = hoje.strftime("%d-%m-%Y")

email = os.getenv("EMAIL")
password = os.getenv("PASSWORD")


pasta_downloads = r"C:\Users\sigab\Downloads"
pasta_destino = f"C:\\Users\\sigab\\OneDrive\\Driver - BPO\\Caneca Fina\\Fechamentos\\{mes_atual}\\dia a dia"
pasta_base = r"C:\Users\sigab\OneDrive\Driver - BPO\Caneca Fina\Fechamentos"
caminho_outra_planilha = fr"C:\Users\sigab\OneDrive\Driver - BPO\Caneca Fina\Fechamentos\{mes_atual}\dia a dia\{data_anterior}.xlsx"
caminho_arquivo = r"C:\Users\sigab\OneDrive\Driver - BPO\Caneca Fina\Fechamentos\12-2024\Vendas_Atualizado.xlsx"

arquivo_original = os.path.join(pasta_downloads, "Listagem.xlsx")
#arquivo_novo = os.path.join(pasta_destino, f"{data}.xlsx")

def configurar_driver():
    edge_options = webdriver.EdgeOptions()
    prefs = {
        "download.default_directory": r"C:\Users\sigab\Downloads", 
        "download.prompt_for_download": False, 
        "download.directory_upgrade": True,
        "safebrowsing.enabled": True
    }
    edge_options.add_experimental_option("prefs", prefs)
    edge_options.add_argument("--start-maximized")
    
    driver = webdriver.Edge(service=Service(EdgeChromiumDriverManager().install()), options=edge_options)
    return driver

def realizar_login(driver):
    driver.get("https://portal.netcontroll.com.br/#/auth/login")


    time.sleep(3)

    email_field = driver.find_element(By.ID, "email")  
    email_field.send_keys(email) 

    password_field = driver.find_element(By.ID, "password")  
    password_field.send_keys(password)  



    login_button = driver.find_element(By.XPATH, "//button[contains(text(), 'Login')]")
    login_button.click()


    time.sleep(5)

def navegar_para_relatorios(driver):

    relatorios_link = driver.find_element(By.XPATH, "//a[@routerlink='/relatorio']")


    relatorios_link.click()


    driver.implicitly_wait(10)

def criar_pasta(mes_anterior, mes_atual):

    if mes_anterior != mes_atual:
        
        nova_pasta = os.path.join(pasta_base, mes_atual, "dia a dia")

        if not os.path.exists(nova_pasta):
            os.makedirs(nova_pasta)
            print(f"Pasta criada: {nova_pasta}")
        else:
            print("A pasta já existe.")


def verificar_fim_de_semana(data_atual):

    data_obj = datetime.strptime(data_atual, "%d-%m-%Y")
    print(f"Data fornecida: {data_atual}, Data convertida: {data_obj}")

    data_anterior = (data_obj - timedelta(days=1)).strftime("%d-%m-%Y")
    print(f"Data anterior: {data_anterior}")
    
    if data_obj.weekday() == 0:  
        sexta = (data_obj - timedelta(days=3)).strftime("%d-%m-%Y")
        sabado = (data_obj - timedelta(days=2)).strftime("%d-%m-%Y")
        print(f"Segunda-feira detectada, retornando sexta e sábado: {sexta}, {sabado}")
        return [sexta, sabado, data_anterior]
    else:
        return [data_anterior]


def gerar_relatorio_vendas(driver):

    vendas_link = driver.find_element(By.XPATH, "//a[contains(text(), '01 - Vendas')]")


    vendas_link.click()


    driver.implicitly_wait(10)

    produtos_link = driver.find_element(By.XPATH, "//a[contains(text(), '101-Produtos')]")


    produtos_link.click()


    driver.implicitly_wait(10)

    checkbox1 = driver.find_element(By.XPATH, "//label[contains(text(), 'Exibir preço e custo (Atual)')]")


    if checkbox1.is_displayed() and checkbox1.is_enabled():
        checkbox1.click()

    checkbox2 = driver.find_element(By.XPATH, "//label[contains(text(), 'Exibir Nº Caixa')]")

    if checkbox2.is_displayed() and checkbox2.is_enabled():
        checkbox2.click()

    checkbox3 = driver.find_element(By.XPATH, "//label[contains(text(), 'Exibir Tipo Emissor Fiscal')]")

    if checkbox3.is_displayed() and checkbox3.is_enabled():
        checkbox3.click()

def selecionar_data(data):

    input_data_inicial = driver.find_element(By.XPATH, "(//dx-date-box//input[@class='dx-texteditor-input'])[1]")
    input_data_inicial.clear()
    input_data_inicial.send_keys(data) 

    input_data_final = driver.find_element(By.XPATH, "(//dx-date-box//input[@class='dx-texteditor-input'])[2]")
    input_data_final.clear()
    input_data_final.send_keys(data)

    botao_excel = driver.find_element(By.XPATH, "//div[@class='dx-button-content']//i[@class='dx-icon dx-icon-export-excel-button']")


    botao_excel.click()

    time.sleep(8)

def obter_pasta_destino(data, mes_atual, mes_anterior):
   
    mes_data = datetime.strptime(data, "%d-%m-%Y").strftime("%m-%Y")

    print(f"Data fornecida: {data}")
    print(f"mes_atual: {mes_atual}")
    print(f"mes_anterior: {mes_anterior}")
    print(f"mes_data: {mes_data}")


    if mes_data == mes_atual:
        pasta_destino = f"C:\\Users\\sigab\\OneDrive\\Driver - BPO\\Caneca Fina\\Fechamentos\\{mes_atual}\\dia a dia"
        caminho_outra_planilha = fr"C:\Users\sigab\OneDrive\Driver - BPO\Caneca Fina\Fechamentos\{mes_atual}\dia a dia\{data}.xlsx"
    else:
        pasta_destino = f"C:\\Users\\sigab\\OneDrive\\Driver - BPO\\Caneca Fina\\Fechamentos\\{mes_anterior}\\dia a dia"
        caminho_outra_planilha = fr"C:\Users\sigab\OneDrive\Driver - BPO\Caneca Fina\Fechamentos\{mes_anterior}\dia a dia\{data}.xlsx"

    return pasta_destino, caminho_outra_planilha

def criar_excel(data, pasta_destino, arquivo_original):

    global caminho_outra_planilha

    if os.path.exists(arquivo_original):
        print("Arquivo encontrado!")
        df = pd.read_excel(arquivo_original)
        df.insert(0, "Data", data)
        df = df[df["Código"].astype(str).str.isnumeric()]
        df = df.dropna(subset=['Nome']) 

        print(df)

        pasta_destino, caminho_outra_planilha = obter_pasta_destino(data, mes_atual, mes_anterior)

        arquivo_novo = os.path.join(pasta_destino, f"{data}.xlsx")
        
        df.to_excel(arquivo_novo, index=False)

        
    
    else:
        print("Arquivo NÃO encontrado. Verifique o caminho.")

    time.sleep(5)

def deletar_arquivo():
    if os.path.exists(arquivo_original):
        os.remove(arquivo_original)
        print("Arquivo Listagem.xlsx deletado com sucesso.")
    else:
        print("O arquivo Listagem.xlsx não foi encontrado.")

def editar_planilhas(caminho_outra_planilha):

    global planilha
    global vendas
    global df_outra_planilha
    global ultima_linha_df
    global pasta_destino

    pasta_destino, caminho_outra_planilha = obter_pasta_destino(data, mes_atual, mes_anterior)
    print(caminho_outra_planilha)
    print("aaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaa")

    print("Carregando planilhas...")

    planilha = pd.read_excel(caminho_arquivo, sheet_name=None)
    print(f"Planilhas carregadas: {list(planilha.keys())}")

    vendas = planilha['Sheet1']
    print(f"Colunas da aba 'Vendas': {vendas.columns}")

    df_outra_planilha = pd.read_excel(caminho_outra_planilha, sheet_name='Sheet1')
    print(f"Colunas da segunda planilha antes do renomeio: {df_outra_planilha.columns}")

    print("Exemplo de datas convertidas em 'Vendas':")
    print(vendas[['Data Venda']].head())

    print("Exemplo de datas convertidas na segunda planilha:")
    print(df_outra_planilha[['Data']].head())
 
    df_outra_planilha = df_outra_planilha.rename(columns={'Data': 'Data Venda', 'Valor Total': 'Valor', 'Emissor Fiscal': 'Tipo Cfe'})

    df_outra_planilha["Data Venda"] = df_outra_planilha["Data Venda"].astype(str).str.replace("-", "/")

    print(f"Colunas da segunda planilha após renomeio: {df_outra_planilha.columns}")

    ultima_linha_df = vendas['Data Venda'].last_valid_index() + 1
    print(f"Última linha válida na aba 'Vendas': {ultima_linha_df}")


def concatenar_planilhas(caminho_arquivo, vendas, df_outra_planilha, ultima_linha_df):

    df_concatenado = pd.concat([vendas.iloc[: ultima_linha_df], df_outra_planilha], ignore_index=True)

    print(f"Tamanho do dataframe antes: {len(vendas)}, tamanho do dataframe depois da concatenação: {len(df_concatenado)}")

    wb = load_workbook(caminho_arquivo, data_only=False)

    ws = wb["Sheet1"] 

    for i, row in df_concatenado.iterrows():
        for j, value in enumerate(row):
            ws.cell(row=i+2, column=j+1, value=value)


    time.sleep(60)

    for row in range(2, len(df_concatenado) + 2): 
        ws[f'L{row}'] = f'=VLOOKUP(C{row},Cardapio!B:J,8,0)'
        ws[f'M{row}'] = f'=VLOOKUP(C{row},Cardapio!B:J,9,0)'

    time.sleep(60)

    wb.save(caminho_arquivo)

    

driver = configurar_driver()
realizar_login(driver)
navegar_para_relatorios(driver)

datas_para_processar = verificar_fim_de_semana(data_atual)
gerar_relatorio_vendas(driver)

for data in datas_para_processar:
    print(f"Iniciando o processamento para a data: {data}")
   
    selecionar_data(data)
    criar_pasta(data, mes_atual)  
    criar_excel(data, pasta_destino, arquivo_original)
    deletar_arquivo() 
    editar_planilhas(caminho_outra_planilha)  
    concatenar_planilhas(caminho_arquivo,vendas, df_outra_planilha, ultima_linha_df)

print("Processo concluído.")


driver.quit()